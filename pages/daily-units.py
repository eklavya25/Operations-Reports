import streamlit as st
import pandas as pd
import re
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook

st.set_page_config(page_title="Daily Units", layout="wide")

st.title("📊 Daily Units")

st.info("Steps")
st.write("Step-1 Download Booked data")
st.write("Step-2 Remove top 6 rows and the last row")
st.write("Step-3 Find all the Denture cases from EasyDentConnect")
st.write("Step-4 Save the file as .xlsx")
st.write("Step-5 Upload the .xlsx file here")
# Step 1: Upload file
uploaded_file = st.file_uploader("📂 Upload the Booked Data after finding all the Dentures (Remove top 6 rows and the last row)", type=["xlsx"])

uploaded_file_2 = st.file_uploader("📂 Upload the Daily Units File here", type=["xlsx"])

# Step 2: Date input
user_date_input = st.text_input("📅 Enter cutoff date (DD/MM)", placeholder="e.g. 20/10")


user_limit = st.number_input("How many 3shape units are in EDDL?", min_value=0, step=50)

if uploaded_file and user_date_input and uploaded_file_2:
    try:
        cutoff_date = datetime.strptime(user_date_input, "%d/%m").replace(year=2025)
        cutoff_time = datetime.strptime("05:30", "%H:%M").time()

        
        df1 = pd.read_excel(uploaded_file, engine='openpyxl')
        # df1 = df.iloc[6:-1].reset_index(drop=True)
        # df1.columns = df1.iloc[0]
        # df1 = df1.drop(df1.index[0])

        df1['#Units'] = pd.to_numeric(df1['#Units'], errors='coerce').fillna(0)
        df1.loc[df1['#Units'] == 0, '#Units'] = 1
        df1.loc[df1['Case Type'].isin(['M', 'M+']), '#Units'] = (
            df1.loc[df1['Case Type'].isin(['M', 'M+']), '#Units']
            .apply(lambda x: x if x in [1, 2] else (2 if x > 20 else 1))
        )

        df1['Order ID'] = df1['Order ID'].astype(str).apply(
            lambda x: re.search(r'"(\d+)"', x).group(1) if re.search(r'"(\d+)"', x) else x
        )
        df1['Order ID'] = pd.to_numeric(df1['Order ID'], errors='ignore')

        Redesign = df1[df1['Order ID'].astype(str).str.contains('r', case=False, na=False)]
        Redesign_count = len(Redesign)
        Redesign_sum = Redesign['#Units'].sum()

        Restarted = df1[df1['Restart Date'].notna() & (df1['Restart Date'].astype(str).str.strip() != '')]
        Restarted_count = len(Restarted)
        Restarted_sum = Restarted['#Units'].sum()

        df1 = df1[~df1.index.isin(Redesign.index)]

        df1['Case In'] = pd.to_datetime(df1['Case In'], errors='coerce')
        df1 = df1.sort_values(by='Case In', ascending=True).reset_index(drop=True)
        df1['date'] = df1['Case In'].dt.strftime('%d-%m')
        df1['time'] = df1['Case In'].dt.strftime('%H:%M')

        df1['date_dt'] = pd.to_datetime(df1['date'] + '-2025', format='%d-%m-%Y', errors='coerce')
        df1['time_dt'] = pd.to_datetime(df1['time'], format='%H:%M', errors='coerce').dt.time

        df1 = df1[
            (df1['date_dt'] > cutoff_date) |
            ((df1['date_dt'] == cutoff_date) & (df1['time_dt'] >= cutoff_time))
        ].reset_index(drop=True)

        df1 = df1.drop(columns=['date_dt', 'time_dt'])


        # Example condition
        mask = (
            (df1["Destination"] == "Easydent Dental Lab") &
            (df1["Software"].isna()) &
            (
                df1["Lab Name"].str.contains("EDDL Impression|Showcase Dental Lab|4G Dental Lab", case=False, na=False)
            )
        )

        # Update the blank values to "Exocad"
        df1.loc[mask, "Software"] = "Exocad"



        # Step 1: Filter matching rows
        mask = (
            (df1["Destination"] == "Easydent Dental Lab") &
            (df1["Software"].isna()) &
            (df1["Lab Name"].str.contains("Easy Dent Dental Lab", case=False, na=False))
        )

        filtered = df1[mask].copy()

        # Step 2: Cumulative sum
        filtered["cum_units"] = filtered["#Units"].cumsum()

        # Step 3: Assign Software based on cumulative sum
        filtered["Software"] = filtered["cum_units"].apply(lambda x: "3Shape" if x <= user_limit else "Exocad")

        # Step 4: Update original dataframe
        df1.loc[filtered.index, "Software"] = filtered["Software"]


        Exocad_software = df1[df1["Software"] == "Exocad"]
        Exocad_cases = Exocad_software["#Units"].count() 
        Exocad_units = Exocad_software["#Units"].sum()
        
        # Step 1: create a modified copy for aggregation
        df1_mod = df1.copy()

        # Step 2: create adjusted Lab Name column based on Destination
        df1_mod['Lab Name Adjusted'] = df1_mod.apply(
            lambda x: f"{x['Lab Name']}- EDDL" if x['Destination'] == 'Easydent Dental Lab' else x['Lab Name'],
            axis=1
        )

        # Step 3: aggregate normally on adjusted lab names
        summary_df = (
            df1_mod.groupby('Lab Name Adjusted', as_index=False)
            .agg(
                First_Order_ID=('Order ID', 'min'),
                Last_Order_ID=('Order ID', 'max'),
                Count=('Order ID', 'count'),
                Sum=('#Units', 'sum')
            )
        )

        # Step 4: Hold and Cancel summaries
        hold_sum = (
            df1_mod[df1_mod['Hold'] == 'Y']
            .groupby('Lab Name Adjusted')['#Units']
            .sum()
            .rename('Hold')
        )

        cancel_sum = (
            df1_mod[df1_mod['Cancel'] == 'Y']
            .groupby('Lab Name Adjusted')['#Units']
            .sum()
            .rename('Cancel')
        )

        # Step 5: merge Hold and Cancel into summary
        summary_df = (
            summary_df
            .merge(hold_sum, on='Lab Name Adjusted', how='left')
            .merge(cancel_sum, on='Lab Name Adjusted', how='left')
        )

        # Step 6: fill missing values with 0
        summary_df[['Hold', 'Cancel']] = summary_df[['Hold', 'Cancel']].fillna(0)

        # ✅ Final rename (optional)
        summary_df = summary_df.rename(columns={'Lab Name Adjusted': 'Lab Name'})




        labs_to_clean = [
            "Easy Dent Dental Lab",
            "4G Dental Lab",
            "Showcase Dental Lab",
            "EDDL Implants",
            "Dental Infinity Laboratory Ltd",
            "EDDL Impression",
            "Marvel Dental"
        ]

        def clean_if_target(name):
            if not isinstance(name, str):
                return name
            if any(lab.lower() in name.lower() for lab in labs_to_clean):
                cleaned = re.sub(r'^-+', '', name)
                cleaned = re.sub(r'[\s-]*EDDL[\s-]*$', '', cleaned, flags=re.IGNORECASE)
                return cleaned.strip()
            return name

        summary_df["Lab Name"] = summary_df["Lab Name"].apply(clean_if_target)


        
        df_x = summary_df.copy()

        base_labs = df_x["Lab Name"].str.extract(r"^(.*?)-\s*EDDL$", expand=False).dropna().unique()

        for base in base_labs:
            mask = df_x["Lab Name"].isin([base, f"{base}-EDDL", f"{base}- EDDL"])
            if mask.sum() > 0:
                min_first = df_x.loc[mask, "First_Order_ID"].min()
                max_last = df_x.loc[mask, "Last_Order_ID"].max()
                df_x.loc[mask, "First_Order_ID"] = min_first
                df_x.loc[mask, "Last_Order_ID"] = max_last

        summary_df = df_x

        # --- Aggregate summary_df to handle duplicates ---
        summary_df_agg = summary_df.groupby("Lab Name", as_index=False).agg(
        First_Order_ID=("First_Order_ID", "min"),
        Last_Order_ID=("Last_Order_ID", "max"),
        Count=("Count", "sum"),
        Sum=("Sum", "sum"),
        Hold=("Hold", "sum"),
        Cancel=("Cancel", "sum")
        )



        Total_Cases = sum(summary_df["Count"])
        Total_Units = sum(summary_df["Sum"])
        Total_Hold = sum(summary_df["Hold"])
        Total_Cancel = sum(summary_df["Cancel"])


        wb = load_workbook(uploaded_file_2)
        source = wb["format"]

        # Create new sheet with the same formatting and formulas
        target = wb.copy_worksheet(source)
        target.title = "Todays Units"

        # --- Step 3: Fill data into 'Todays Units' ---
        for row in range(1, target.max_row + 1):
            lab_name = target[f"A{row}"].value
            if lab_name:
                match = summary_df_agg[summary_df_agg["Lab Name"] == lab_name]
                if not match.empty:
                    match = match.iloc[0]
                    target[f"B{row}"].value = match["First_Order_ID"]
                    target[f"C{row}"].value = match["Last_Order_ID"]
                    target[f"F{row}"].value = match["Count"]
                    target[f"G{row}"].value = match["Sum"]
                    target[f"I{row}"].value = match["Hold"]
                    target[f"J{row}"].value = match["Cancel"]
                elif lab_name == "Total No. of Exo (Units)":
                    target[f"B{row}"].value = Exocad_cases
                    target[f"C{row}"].value = Exocad_units

                elif lab_name == "Re-Design":
                    target[f"B{row}"].value = Redesign_count
                    target[f"C{row}"].value = Redesign_sum

                elif lab_name == "Restarted":
                    target[f"B{row}"].value = Restarted_count
                    target[f"C{row}"].value = Restarted_sum

        # --- Step 4: Save workbook into memory for download ---
        final_buffer = BytesIO()
        wb.save(final_buffer)
        final_buffer.seek(0)

        # # --- Step 5: Streamlit download button ---
        # st.download_button(
        #     label="📘 Download The Full Report",
        #     data=final_buffer,
        #     file_name="Full_Report.xlsx",
        #     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        # )



        st.subheader("📄 Preview")
        st.dataframe(summary_df, use_container_width=True)

        st.subheader("Redesign and Restart Cases")
        st.write(f"**Redesign Cases:** {Redesign_count} | **Units:** {Redesign_sum}")
        st.write(f"**Restart Cases:** {Restarted_count} | **Units:** {Restarted_sum}")

        st.subheader("Summary")

        st.write(f"**Total Cases:** {Total_Cases}")
        st.write(f"**Total Units:** {Total_Units}")
        st.write(f"**Total Hold:** {int (Total_Hold)}")
        st.write(f"**Total Cancel:** {int (Total_Cancel)}")

        st.subheader("Total cases and units in Exocad")
        st.write(f"**Exocad cases:** {int (Exocad_cases)}")
        st.write(f"**Exocad Units:** {int (Exocad_units)}")

        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            summary_df.to_excel(writer, index=False, sheet_name='Summary')

        # Move buffer position to the beginning
        buffer.seek(0)

        # Download button
        st.download_button(
            label="📥 Download Summary Excel",
            data=buffer,
            file_name="Lab_Summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # --- Step 5: Streamlit download button ---
        st.download_button(
            label="📘 Download The Full Report",
            data=final_buffer,
            file_name="Full_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"❌ Error: {e}")

# else:
#     st.info("👆 Please upload your `.xlsx` file and enter a cutoff date (DD/MM) to begin.")
